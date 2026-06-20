"""
Local knowledge base / RAG pipeline.

ingest:  save record → detect type → parse (txt/md/csv/json/xlsx/pdf) → chunk →
         local embed (Ollama) → store chunks + embeddings.
retrieve: embed query → cosine over stored local embeddings → trust/lifecycle
         filtering → grounded sources.

Vector store: embeddings are stored locally in Postgres (document_chunks.
embedding_json) and ranked with in-process cosine similarity. This is fully
local and needs no extra service. pgvector is a documented optional upgrade.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

from sqlalchemy import select

from config import settings
from db.models import Document, DocumentChunk
from services.embedding import embed_text, embed_texts, cosine

# NOTE: legacy binary .xls (BIFF) is intentionally NOT supported — openpyxl only
# reads the OOXML .xlsx format and would raise on a real .xls. Convert to .xlsx.
SUPPORTED = {"txt", "md", "csv", "json", "xlsx", "pdf", "yaml", "yml"}
CHUNK_CHARS = 900
CHUNK_OVERLAP = 120

# Parser resource ceilings — a hostile/corrupt file (decompression bomb, huge
# sheet) must not exhaust memory. Anything beyond the cap is simply not indexed.
MAX_PDF_PAGES = 1000
MAX_SHEET_ROWS = 10_000
MAX_CSV_ROWS = 50_000

# trust ordering for guidance
TRUST_ORDER = ["untrusted", "low", "medium", "high", "authoritative"]
# document_status values that must NOT be auto-retrieved
EXCLUDED_STATUS = {"deprecated", "archived"}


# ── parsing ──────────────────────────────────────────────────────────
def detect_tabular_schema(rows: List[List[str]]) -> Dict:
    if not rows:
        return {"columns": [], "row_count": 0, "missing_value_summary": {}, "sample_rows": []}
    header = rows[0]
    body = rows[1:]
    missing = {col: 0 for col in header}
    for r in body:
        for i, col in enumerate(header):
            if i >= len(r) or str(r[i]).strip() == "":
                missing[col] += 1
    return {
        "columns": header,
        "column_count": len(header),
        "row_count": len(body),
        "missing_value_summary": missing,
        "sample_rows": body[:5],
    }


def parse_file(abs_path: str, ext: str) -> Tuple[List[Dict], Dict]:
    """Return (segments, extra). segments: [{text, page?, sheet?}]. extra may
    carry a detected tabular schema for csv/xlsx."""
    ext = ext.lower().lstrip(".")
    extra: Dict = {}
    segments: List[Dict] = []

    if ext in {"txt", "md", "yaml", "yml"}:
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            segments.append({"text": f.read()})
    elif ext == "json":
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            raw = f.read()
        try:
            segments.append({"text": json.dumps(json.loads(raw), indent=2, ensure_ascii=False)})
        except Exception:
            segments.append({"text": raw})
    elif ext == "csv":
        import itertools
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            rows = list(itertools.islice(csv.reader(f), MAX_CSV_ROWS))
        schema = detect_tabular_schema(rows)
        extra["schema"] = schema
        text = "Columns: " + ", ".join(schema["columns"]) + f"\nRows: {schema['row_count']}\n"
        text += "\n".join(", ".join(r) for r in rows[:200])
        segments.append({"text": text})
    elif ext == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(abs_path, read_only=True, data_only=True)
            sheet_schemas = {}
            import itertools
            for ws in wb.worksheets:
                rows = [[("" if c is None else str(c)) for c in row]
                        for row in itertools.islice(ws.iter_rows(values_only=True), MAX_SHEET_ROWS)]
                schema = detect_tabular_schema(rows)
                sheet_schemas[ws.title] = schema
                text = f"Sheet: {ws.title}\nColumns: " + ", ".join(schema["columns"]) + \
                       f"\nRows: {schema['row_count']}\n" + \
                       "\n".join(", ".join(r) for r in rows[:200])
                segments.append({"text": text, "sheet": ws.title})
            extra["schema"] = sheet_schemas
        except ImportError:
            raise RuntimeError("openpyxl is required to parse .xlsx — pip install openpyxl")
    elif ext == "pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(abs_path)
            for i, page in enumerate(reader.pages):
                if i >= MAX_PDF_PAGES:
                    break
                segments.append({"text": page.extract_text() or "", "page": i + 1})
        except ImportError:
            raise RuntimeError("pypdf is required to parse .pdf — pip install pypdf")
    else:
        raise RuntimeError(f"Unsupported file type for knowledge ingestion: {ext}")

    return segments, extra


def chunk_segments(segments: List[Dict]) -> List[Dict]:
    chunks: List[Dict] = []
    for seg in segments:
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        start = 0
        n = len(text)
        while start < n:
            end = min(start + CHUNK_CHARS, n)
            # Prefer to break on a whitespace boundary near CHUNK_CHARS so words and
            # multi-byte tokens aren't split mid-character (which degrades embeddings).
            if end < n:
                window = text[start:end]
                brk = max(window.rfind(" "), window.rfind("\n"), window.rfind("\t"))
                if brk >= int(CHUNK_CHARS * 0.6):   # only honour a reasonably-late break
                    end = start + brk + 1
            piece = text[start:end].strip()
            if piece:
                chunks.append({"content": piece, "page": seg.get("page"), "sheet": seg.get("sheet")})
            if end >= n:
                break
            start = max(end - CHUNK_OVERLAP, start + 1)
    return chunks


# ── ingestion ────────────────────────────────────────────────────────
async def ingest_document(db, file_meta: Dict, *, owner: str = None, client_name: str = None,
                          project_name: str = None, source_type: str = None,
                          sensitivity_level: str = "internal", abs_path: str = None) -> Document:
    ext = file_meta["ext"].lower()
    if ext not in SUPPORTED:
        raise RuntimeError(f"Unsupported file type: {ext}")

    doc = Document(
        file_id=file_meta["file_id"], filename=file_meta["original_filename"],
        stored_path=file_meta["stored_path"], sha256=file_meta.get("sha256", ""),
        file_type=ext, size_bytes=file_meta.get("size", 0),
        document_status="raw", sensitivity_level=sensitivity_level,
        trust_level="untrusted", owner=owner, client_name=client_name,
        project_name=project_name, source_type=source_type,
    )
    db.add(doc)
    await db.flush()

    segments, extra = parse_file(abs_path or file_meta["abs_path"], ext)
    doc.document_status = "parsed"
    if extra.get("schema"):
        doc.metadata_json = {**(doc.metadata_json or {}), "schema": extra["schema"]}

    chunks = chunk_segments(segments)
    texts = [c["content"] for c in chunks]
    vectors = await embed_texts(texts) if texts else []
    _assert_embedding_complete(chunks, vectors)

    for idx, (c, vec) in enumerate(zip(chunks, vectors)):
        db.add(DocumentChunk(
            document_id=doc.id, chunk_index=idx, content=c["content"],
            embedding_json=vec, token_estimate=max(0, len(c["content"]) // 4),
            page=c.get("page"), sheet=c.get("sheet"),
        ))
    doc.chunk_count = len(chunks)
    doc.document_status = "indexed"  # new docs default: indexed + untrusted
    doc.metadata_json = {**(doc.metadata_json or {}),
                         "embedding_model": settings.OLLAMA_MODEL_EMBEDDING}
    await db.flush()
    return doc


async def reindex_document(db, doc: Document, abs_path: str) -> Document:
    await db.execute(
        DocumentChunk.__table__.delete().where(DocumentChunk.document_id == doc.id)
    )
    segments, extra = parse_file(abs_path, doc.file_type)
    chunks = chunk_segments(segments)
    vectors = await embed_texts([c["content"] for c in chunks]) if chunks else []
    _assert_embedding_complete(chunks, vectors)
    for idx, (c, vec) in enumerate(zip(chunks, vectors)):
        db.add(DocumentChunk(document_id=doc.id, chunk_index=idx, content=c["content"],
                             embedding_json=vec, token_estimate=max(0, len(c["content"]) // 4),
                             page=c.get("page"), sheet=c.get("sheet")))
    doc.chunk_count = len(chunks)
    doc.document_status = "indexed"
    doc.metadata_json = {**(doc.metadata_json or {}),
                         "embedding_model": settings.OLLAMA_MODEL_EMBEDDING}
    await db.flush()
    return doc


def _assert_embedding_complete(chunks, vectors) -> None:
    """A partial embedding batch must fail the whole ingest (transaction rolls
    back) — otherwise zip() silently drops the tail chunks and they become
    permanently unretrievable with no error anywhere."""
    if len(vectors) != len(chunks):
        raise RuntimeError(
            f"Embedding returned {len(vectors)} vectors for {len(chunks)} chunks — "
            "aborting ingest instead of silently indexing a partial document.")


# ── retrieval / grounding ────────────────────────────────────────────
async def retrieve_context(db, query: str, top_k: int = None, filters: Optional[Dict] = None) -> List[Dict]:
    # Clamp top_k to a hard server-side bound so a caller can never force the whole
    # corpus (or an arbitrarily large response) into one prompt/response.
    requested = int(top_k or settings.MAX_CONTEXT_CHUNKS)
    top_k = max(1, min(requested, settings.MAX_RETRIEVAL_TOP_K))
    qvec = await embed_text(query)
    now = datetime.utcnow()

    # Push the cheap, known exclusions into SQL so deprecated/archived chunks are
    # never even loaded into memory (and the scan grows with candidates, not corpus).
    stmt = (select(DocumentChunk, Document)
            .join(Document, Document.id == DocumentChunk.document_id)
            .where(Document.document_status.notin_(EXCLUDED_STATUS)))
    if filters and filters.get("client_name"):
        stmt = stmt.where(Document.client_name == filters["client_name"])
    res = await db.execute(stmt)

    scored = []
    stale_model_docs: set = set()
    for chunk, doc in res.all():
        if _is_expired(doc, now):                    # never auto-use expired knowledge
            continue
        # Embedding-model tracking: vectors from a different embedding model are
        # incomparable (cosine is meaningless/0), so the doc would just silently
        # vanish from retrieval. Skip it AND surface the reason loudly.
        emb_model = (doc.metadata_json or {}).get("embedding_model")
        if emb_model and emb_model != settings.OLLAMA_MODEL_EMBEDDING:
            stale_model_docs.add(str(doc.id))
            continue
        if filters:
            if filters.get("min_trust") and _trust_rank(doc.trust_level) < _trust_rank(filters["min_trust"]):
                continue
            # don't retrieve the in-progress conversation's own ingested transcript
            if filters.get("exclude_conversation_id") and doc.conversation_id is not None \
                    and str(doc.conversation_id) == str(filters["exclude_conversation_id"]):
                continue
        rel = cosine(qvec, chunk.embedding_json or [])
        if rel < settings.MIN_RELEVANCE:             # drop irrelevant / dimension-mismatched chunks
            continue
        scored.append((rel, chunk, doc))
    scored.sort(key=lambda x: x[0], reverse=True)
    if stale_model_docs:
        print(f"⚠️  Knowledge: {len(stale_model_docs)} document(s) were indexed with a different "
              f"embedding model than the current '{settings.OLLAMA_MODEL_EMBEDDING}' and were "
              "excluded from retrieval. Reindex them (POST /api/knowledge/documents/{id}/reindex).")

    out = []
    for rel, chunk, doc in scored[:top_k]:
        out.append({
            "document_id": str(doc.id), "chunk_id": str(chunk.id), "filename": doc.filename,
            # `content` is the full chunk used as the grounding payload; `content_preview`
            # is the short string for UI lists. (Previously only the 400-char preview was
            # injected, so the model saw <45% of each chunk.)
            "content": chunk.content[:settings.RETRIEVAL_CHUNK_CHARS],
            "content_preview": chunk.content[:400], "relevance": round(float(rel), 4),
            "relevance_score": round(float(rel), 4),
            "trust_level": doc.trust_level, "document_status": doc.document_status,
            "sensitivity_level": doc.sensitivity_level, "page": chunk.page, "sheet": chunk.sheet,
            "warning": ("low-trust source — use with caution" if doc.trust_level in ("untrusted", "low") else None),
            "metadata": {"client_name": doc.client_name, "project_name": doc.project_name},
        })
    return out


def _is_expired(doc, now: datetime) -> bool:
    exp = getattr(doc, "expiration_date", None)
    if not exp:
        return False
    try:
        if isinstance(exp, datetime):
            return exp < now
        if isinstance(exp, date):
            return exp < now.date()
    except Exception:
        return False
    return False


def format_grounding(sources: List[Dict]) -> str:
    if not sources:
        return "No local knowledge source was used. This answer is based on model reasoning only."
    lines = ["Sources used:"]
    for i, s in enumerate(sources, 1):
        loc = f"sheet: {s['sheet']}" if s.get("sheet") else (f"page: {s['page']}" if s.get("page") else "")
        loc = f" | {loc}" if loc else ""
        lines.append(f"{i}. {s['filename']}{loc} | chunk_id: {s['chunk_id'][:8]} | "
                     f"relevance: {s['relevance']} | trust: {s['trust_level']}")
    return "\n".join(lines)


def _trust_rank(level: str) -> int:
    try:
        return TRUST_ORDER.index(level)
    except ValueError:
        return 0
