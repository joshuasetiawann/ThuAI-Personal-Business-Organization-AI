"""
Local knowledge base / RAG pipeline.

ingest:  save record → detect type → parse (txt/md/csv/json/xlsx/pdf) → chunk →
         local embed (Ollama) → store chunks + embeddings.
retrieve: embed query → cosine over stored local embeddings → trust/lifecycle
         filtering → grounded sources.

Vector store: embeddings are stored locally in Postgres (document_chunks.
embedding_json) and ranked with in-process cosine similarity. This is fully
local and needs no extra service. pgvector is a documented optional upgrade.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Dict, List, Optional, Tuple

from sqlalchemy import select

from config import settings
from db.models import Document, DocumentChunk
from services.embedding import embed_text, embed_texts, cosine

SUPPORTED = {"txt", "md", "csv", "json", "xlsx", "xls", "pdf", "yaml", "yml"}
CHUNK_CHARS = 900
CHUNK_OVERLAP = 120

# trust ordering for guidance
TRUST_ORDER = ["untrusted", "low", "medium", "high", "authoritative"]
# document_status values that must NOT be auto-retrieved
EXCLUDED_STATUS = {"deprecated", "archived"}


# ── parsing ──────────────────────────────────────────────────────────
def detect_tabular_schema(rows: List[List[str]]) -> Dict:
    if not rows:
        return {"columns": [], "row_count": 0, "missing_value_summary": {}, "sample_rows": []}
    header = rows[0]
    body = rows[1:]
    missing = {col: 0 for col in header}
    for r in body:
        for i, col in enumerate(header):
            if i >= len(r) or str(r[i]).strip() == "":
                missing[col] += 1
    return {
        "columns": header,
        "column_count": len(header),
        "row_count": len(body),
        "missing_value_summary": missing,
        "sample_rows": body[:5],
    }


def parse_file(abs_path: str, ext: str) -> Tuple[List[Dict], Dict]:
    """Return (segments, extra). segments: [{text, page?, sheet?}]. extra may
    carry a detected tabular schema for csv/xlsx."""
    ext = ext.lower().lstrip(".")
    extra: Dict = {}
    segments: List[Dict] = []

    if ext in {"txt", "md", "yaml", "yml"}:
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            segments.append({"text": f.read()})
    elif ext == "json":
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            raw = f.read()
        try:
            segments.append({"text": json.dumps(json.loads(raw), indent=2, ensure_ascii=False)})
        except Exception:
            segments.append({"text": raw})
    elif ext == "csv":
        with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
            rows = list(csv.reader(f))
        schema = detect_tabular_schema(rows)
        extra["schema"] = schema
        text = "Columns: " + ", ".join(schema["columns"]) + f"\nRows: {schema['row_count']}\n"
        text += "\n".join(", ".join(r) for r in rows[:200])
        segments.append({"text": text})
    elif ext in {"xlsx", "xls"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(abs_path, read_only=True, data_only=True)
            sheet_schemas = {}
            for ws in wb.worksheets:
                rows = [[("" if c is None else str(c)) for c in row]
                        for row in ws.iter_rows(values_only=True)]
                schema = detect_tabular_schema(rows)
                sheet_schemas[ws.title] = schema
                text = f"Sheet: {ws.title}\nColumns: " + ", ".join(schema["columns"]) + \
                       f"\nRows: {schema['row_count']}\n" + \
                       "\n".join(", ".join(r) for r in rows[:200])
                segments.append({"text": text, "sheet": ws.title})
            extra["schema"] = sheet_schemas
        except ImportError:
            raise RuntimeError("openpyxl is required to parse .xlsx — pip install openpyxl")
    elif ext == "pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(abs_path)
            for i, page in enumerate(reader.pages):
                segments.append({"text": page.extract_text() or "", "page": i + 1})
        except ImportError:
            raise RuntimeError("pypdf is required to parse .pdf — pip install pypdf")
    else:
        raise RuntimeError(f"Unsupported file type for knowledge ingestion: {ext}")

    return segments, extra


def chunk_segments(segments: List[Dict]) -> List[Dict]:
    chunks: List[Dict] = []
    for seg in segments:
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        start = 0
        while start < len(text):
            piece = text[start:start + CHUNK_CHARS]
            chunks.append({"content": piece, "page": seg.get("page"), "sheet": seg.get("sheet")})
            if start + CHUNK_CHARS >= len(text):
                break
            start += CHUNK_CHARS - CHUNK_OVERLAP
    return chunks


# ── ingestion ────────────────────────────────────────────────────────
async def ingest_document(db, file_meta: Dict, *, owner: str = None, client_name: str = None,
                          project_name: str = None, source_type: str = None,
                          sensitivity_level: str = "internal", abs_path: str = None) -> Document:
    ext = file_meta["ext"].lower()
    if ext not in SUPPORTED:
        raise RuntimeError(f"Unsupported file type: {ext}")

    doc = Document(
        file_id=file_meta["file_id"], filename=file_meta["original_filename"],
        stored_path=file_meta["stored_path"], sha256=file_meta.get("sha256", ""),
        file_type=ext, size_bytes=file_meta.get("size", 0),
        document_status="raw", sensitivity_level=sensitivity_level,
        trust_level="untrusted", owner=owner, client_name=client_name,
        project_name=project_name, source_type=source_type,
    )
    db.add(doc)
    await db.flush()

    segments, extra = parse_file(abs_path or file_meta["abs_path"], ext)
    doc.document_status = "parsed"
    if extra.get("schema"):
        doc.metadata_json = {**(doc.metadata_json or {}), "schema": extra["schema"]}

    chunks = chunk_segments(segments)
    texts = [c["content"] for c in chunks]
    vectors = await embed_texts(texts) if texts else []

    for idx, (c, vec) in enumerate(zip(chunks, vectors)):
        db.add(DocumentChunk(
            document_id=doc.id, chunk_index=idx, content=c["content"],
            embedding_json=vec, token_estimate=max(0, len(c["content"]) // 4),
            page=c.get("page"), sheet=c.get("sheet"),
        ))
    doc.chunk_count = len(chunks)
    doc.document_status = "indexed"  # new docs default: indexed + untrusted
    await db.flush()
    return doc


async def reindex_document(db, doc: Document, abs_path: str) -> Document:
    await db.execute(
        DocumentChunk.__table__.delete().where(DocumentChunk.document_id == doc.id)
    )
    segments, extra = parse_file(abs_path, doc.file_type)
    chunks = chunk_segments(segments)
    vectors = await embed_texts([c["content"] for c in chunks]) if chunks else []
    for idx, (c, vec) in enumerate(zip(chunks, vectors)):
        db.add(DocumentChunk(document_id=doc.id, chunk_index=idx, content=c["content"],
                             embedding_json=vec, token_estimate=max(0, len(c["content"]) // 4),
                             page=c.get("page"), sheet=c.get("sheet")))
    doc.chunk_count = len(chunks)
    doc.document_status = "indexed"
    await db.flush()
    return doc


# ── retrieval / grounding ────────────────────────────────────────────
async def retrieve_context(db, query: str, top_k: int = None, filters: Optional[Dict] = None) -> List[Dict]:
    top_k = top_k or settings.MAX_CONTEXT_CHUNKS
    qvec = await embed_text(query)

    res = await db.execute(select(DocumentChunk, Document).join(
        Document, Document.id == DocumentChunk.document_id))
    scored = []
    for chunk, doc in res.all():
        if doc.document_status in EXCLUDED_STATUS:   # never auto-use deprecated/archived
            continue
        if filters:
            if filters.get("client_name") and doc.client_name != filters["client_name"]:
                continue
            if filters.get("min_trust") and _trust_rank(doc.trust_level) < _trust_rank(filters["min_trust"]):
                continue
            # don't retrieve the in-progress conversation's own ingested transcript
            if filters.get("exclude_conversation_id") and doc.conversation_id is not None \
                    and str(doc.conversation_id) == str(filters["exclude_conversation_id"]):
                continue
        rel = cosine(qvec, chunk.embedding_json or [])
        scored.append((rel, chunk, doc))
    scored.sort(key=lambda x: x[0], reverse=True)

    out = []
    for rel, chunk, doc in scored[:top_k]:
        out.append({
            "document_id": str(doc.id), "chunk_id": str(chunk.id), "filename": doc.filename,
            "content_preview": chunk.content[:400], "relevance": round(float(rel), 4),
            "relevance_score": round(float(rel), 4),
            "trust_level": doc.trust_level, "document_status": doc.document_status,
            "sensitivity_level": doc.sensitivity_level, "page": chunk.page, "sheet": chunk.sheet,
            "warning": ("low-trust source — use with caution" if doc.trust_level in ("untrusted", "low") else None),
            "metadata": {"client_name": doc.client_name, "project_name": doc.project_name},
        })
    return out


def format_grounding(sources: List[Dict]) -> str:
    if not sources:
        return "No local knowledge source was used. This answer is based on model reasoning only."
    lines = ["Sources used:"]
    for i, s in enumerate(sources, 1):
        loc = f"sheet: {s['sheet']}" if s.get("sheet") else (f"page: {s['page']}" if s.get("page") else "")
        loc = f" | {loc}" if loc else ""
        lines.append(f"{i}. {s['filename']}{loc} | chunk_id: {s['chunk_id'][:8]} | "
                     f"relevance: {s['relevance']} | trust: {s['trust_level']}")
    return "\n".join(lines)


def _trust_rank(level: str) -> int:
    try:
        return TRUST_ORDER.index(level)
    except ValueError:
        return 0
